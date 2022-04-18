import math

import numpy as np
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.utils.cell import coordinate_from_string
from scipy import interpolate


class WorkbookWrapper:
    def __init__(self):
        self.workbook_path = None
        self.workbook = None
        self.sheet = None

    def open_workbook(self, workbook_path):
        self.workbook_path = workbook_path
        self.workbook = load_workbook(workbook_path)

    def set_sheet(self, sheet_name):
        self.sheet = self.workbook[sheet_name]

    def process_data_set(self, data_set):
        domain_label = data_set.labels[0]
        collections_labels = data_set.labels[1:]
        collections_count = len(data_set.collections_cells[domain_label])

        for label in collections_labels:
            if len(data_set.collections_cells[label]) != collections_count:
                raise ValueError('All series needs to have same length!')

        collections = []

        for collection_n in range(collections_count):
            print('Processing collection: %d of %d' % (collection_n + 1, collections_count))

            collection = CollectionSet()

            start_cell_domain = data_set.collections_cells[domain_label][collection_n]
            print('Loading: %s (base cell %s)' % (domain_label, start_cell_domain))

            domain_values = self.load_single_serie(start_cell_domain)
            collection.add_domain(domain_label, domain_values)

            for label in collections_labels:
                start_cell = data_set.collections_cells[label][collection_n]
                print('Loading: %s (base cell %s)' % (label, start_cell))

                collection_values = self.load_single_serie(start_cell)
                collection.add_collection(label, collection_values)

            for label in collections_labels:
                print('Interpolating: %s(%s)' % (label, domain_label))
                interpolation = interpolate.interp1d(collection.domain_values, collection.get_collection(label),
                                                     bounds_error=False)
                collection.add_interpolation(label, interpolation)

            collections.append(collection)

        print('Create averages')
        average = self.create_average_set(collections)

        print('Saving results')

        for label, start_cell in data_set.averages_cells.items():
            print('Average %s (%s)' % (label, start_cell))
            self.save_values(start_cell, average.get_collection(label))

        for label, start_cell in data_set.deviations_cells.items():
            print('Deviation %s (%s)' % (label, start_cell))
            self.save_values(start_cell, average.get_deviation(label))

    @staticmethod
    def create_average_set(collections):
        average = CollectionSet()
        average.domain_label = collections[0].domain_label
        collections_labels = collections[0].collections.keys()

        x_start = collections[0].domain_values[0]
        x_end = collections[0].domain_values[-1]
        samples = 0

        for collection in collections:
            x_start = min(x_start, collection.domain_values[0])
            x_end = max(x_end, collection.domain_values[-1])
            samples = max(samples, len(collection.domain_values))

        for x in np.linspace(x_start, x_end, num=samples):
            average.domain_values.append(x)

            for label in collections_labels:
                values = [s.interpolation[label](x) for s in collections if not math.isnan(s.interpolation[label](x))]
                v_avg = sum(values) / len(values)
                dev = math.sqrt(sum([(v - v_avg) ** 2 for v in values]) / len(values))
                average.append_collection_value(label, v_avg)
                average.append_deviation(label, dev)

        return average

    def save_values(self, base_cell_str, values):
        column, start_row = self.get_col_and_row(base_cell_str)

        for n in range(len(values)):
            self.sheet.cell(row=start_row + n, column=column).value = values[n]

    def save_workbook(self, workbook_path):
        self.workbook.save(workbook_path)

    def load_single_serie(self, base_cell_str):
        column, start_row = self.get_col_and_row(base_cell_str)

        values = []

        for row in range(start_row, self.sheet.max_row):
            value = self.sheet.cell(column=column, row=row).value
            values.append(value)

        values = [v for v in values if v is not None]
        return np.array(values)

    @staticmethod
    def get_col_and_row(cell_str):
        coordinates = coordinate_from_string(cell_str)
        column = column_index_from_string(coordinates[0])
        row = coordinates[1]
        return column, row


class CollectionSet:
    def __init__(self):
        self.domain_label = None
        self.domain_values = []
        self.collections = {}
        self.interpolation = {}
        self.deviation = {}

    def add_domain(self, label, values):
        self.domain_label = label
        self.domain_values = values

    def add_collection(self, label, values):
        self.collections[label] = values

    def get_collection(self, label):
        return self.domain_values if label == self.domain_label else self.collections[label]

    def append_collection_value(self, label, value):
        if label not in self.collections:
            self.collections[label] = []
        self.collections[label].append(value)

    def add_interpolation(self, label, function):
        if label not in self.collections.keys():
            raise ValueError('No collection with label: %s', label)
        self.interpolation[label] = function

    def append_deviation(self, label, value):
        if label not in self.deviation:
            self.deviation[label] = []
        self.deviation[label].append(value)

    def get_deviation(self, label):
        return self.deviation[label]
