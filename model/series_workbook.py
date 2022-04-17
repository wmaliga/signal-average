import math

import numpy as np
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.utils.cell import coordinate_from_string
from scipy import interpolate


class SeriesWorkbook:
    MAX_SERIES = 5

    def __init__(self):
        self.workbook_path = None
        self.workbook = None
        self.sheet = None

    def open_workbook(self, workbook_path):
        self.workbook_path = workbook_path
        self.workbook = load_workbook(workbook_path)

    def set_sheet(self, sheet_name):
        self.sheet = self.workbook[sheet_name]

    def process_series(self, data_wrapper):
        series_letters_all = [letter for letter in data_wrapper.keys() if len(letter) == 1]
        base_letter = series_letters_all[0]
        series_letters = series_letters_all[1:]
        columns_count = len(data_wrapper[base_letter])

        for letter in series_letters:
            if len(data_wrapper[letter]) != columns_count:
                raise ValueError('All series needs to have same length!')

        series = []

        for series_n in range(columns_count):
            print('Processing column: %d of %d' % (series_n + 1, columns_count))

            serie = Series()

            start_cell = data_wrapper[base_letter][series_n]
            print('Loading: %s (base cell %s)' % (base_letter, start_cell))

            base_values = self.load_single_serie(start_cell)
            serie.add_base(base_letter, base_values)

            for letter in series_letters:
                start_cell = data_wrapper[letter][series_n]
                print('Loading: %s (base cell %s)' % (letter, start_cell))

                serie_values = self.load_single_serie(start_cell)
                serie.add_series(letter, serie_values)

            for letter in series_letters:
                print('Interpolating %s(%s)' % (letter, base_letter))
                interpolation = interpolate.interp1d(serie.base, serie.series[letter], bounds_error=False)
                serie.add_interpolation(letter, interpolation)

            series.append(serie)

        print('Create averages')
        average = self.create_average(series)

        print('Saving results')
        results = [r for r in data_wrapper.keys() if len(r) == 5]

        for result in results:
            start_cell = data_wrapper[result]
            letter, result_type = result.split('_')
            if result_type == 'avg':
                print('Average %s (%s)' % (letter, start_cell))
                self.save_values(start_cell, average.get_series(letter))
            if result_type == "dev":
                print('Deviation %s (%s)' % (letter, start_cell))
                self.save_values(start_cell, average.get_deviation(letter))

    def create_average(self, series):
        average = Series()
        average.base_letter = series[0].base_letter
        series_letters = series[0].series.keys()

        x_start = series[0].base[0]
        x_end = series[0].base[-1]
        samples = 0

        for s in series:
            x_start = min(x_start, s.base[0])
            x_end = max(x_end, s.base[-1])
            samples = max(samples, len(s.base))

        for x in np.linspace(x_start, x_end, num=samples):
            average.base.append(x)

            for letter in series_letters:
                values = [s.interpolation[letter](x) for s in series if not math.isnan(s.interpolation[letter](x))]
                v_avg = sum(values) / len(values)
                dev = math.sqrt(sum([(v - v_avg) ** 2 for v in values]) / len(values))
                average.append_series(letter, v_avg)
                average.append_deviation(letter, dev)

        return average

    def save_values(self, base_cell_str, values):
        column, start_row = self.get_col_and_row(base_cell_str)

        for n in range(len(values)):
            self.sheet.cell(row=start_row + n, column=column).value = values[n]

    def save_workbook(self, workbook_path):
        self.workbook.save(workbook_path)

    def load_single_serie(self, base_cell_str):
        column, start_row = self.get_col_and_row(base_cell_str)

        values = []

        for row in range(start_row, self.sheet.max_row):
            value = self.sheet.cell(column=column, row=row).value
            values.append(value)

        values = [v for v in values if v is not None]
        return np.array(values)

    def get_col_and_row(self, cell_str):
        coordinates = coordinate_from_string(cell_str)
        column = column_index_from_string(coordinates[0])
        row = coordinates[1]
        return column, row

class Series:
    def __init__(self):
        self.base_letter = None
        self.base = []
        self.series = {}
        self.interpolation = {}
        self.deviation = {}
        self.x = []
        self.y = []
        self.xt = None
        self.yt = None
        self.bx = []
        self.by = []

    def add_base(self, letter, values):
        self.base_letter = letter
        self.base = values

    def add_series(self, letter, values):
        self.series[letter] = values

    def get_series(self, letter):
        return self.base if letter == self.base_letter else self.series[letter]

    def append_series(self, letter, value):
        if letter not in self.series:
            self.series[letter] = []
        self.series[letter].append(value)

    def add_interpolation(self, letter, function):
        if letter not in self.series.keys():
            raise ValueError('Missing values for letter: %s', letter)
        self.interpolation[letter] = function

    def append_deviation(self, letter, value):
        if letter not in self.deviation:
            self.deviation[letter] = []
        self.deviation[letter].append(value)

    def get_deviation(self, letter):
        return self.deviation[letter]