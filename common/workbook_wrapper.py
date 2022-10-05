from os.path import splitext

import numpy as np
from numpy import ndarray
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string


class Cell:
    def __init__(self, string: str):
        coordinates = coordinate_from_string(string)
        self.string: str = string
        self.column: int = column_index_from_string(coordinates[0])
        self.row: int = coordinates[1]


class WorkbookWrapper:
    def __init__(self):
        self.workbook_path = None
        self.workbook = None
        self.sheet = None

    def set_sheet(self, sheet_name: str):
        self.sheet = self.workbook[sheet_name]

    def create_sheet(self, name: str):
        if name not in self.workbook:
            self.workbook.create_sheet(name)

    def load_single_series(self, base_cell: Cell) -> ndarray:
        values = []

        for row in range(base_cell.row, self.sheet.max_row):
            value = self.sheet.cell(column=base_cell.column, row=row).value
            values.append(value)

        values = [v for v in values if v is not None]
        return np.array(values)

    def write_single_series(self, base_cell: Cell, values: list, override_sheet=None):
        sheet = self.workbook[override_sheet] if override_sheet else self.sheet

        for n in range(len(values)):
            sheet.cell(row=base_cell.row + n, column=base_cell.column).value = values[n]


def open_workbook(path: str) -> WorkbookWrapper:
    workbook = WorkbookWrapper()
    workbook.workbook_path = path
    workbook.workbook = load_workbook(path)
    return workbook


def save_with_suffix(wrapper: WorkbookWrapper, suffix: str):
    basename, extension = splitext(wrapper.workbook_path)
    path = '{}_{}{}'.format(basename, suffix, extension)
    wrapper.workbook.save(path)
